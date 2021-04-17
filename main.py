import xlsxwriter
import os
from openpyxl import Workbook
import random
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, colors, Fill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

storagelocation = "/Users/weiyushit/OneDrive/Github stuff/NS Detailing/hello_world.xlsx" ##CHANGE EXCEL EXPORT LOCATION
# storagelocation = "hello_world.xlsx" #for mobile compiler

## CHANGE PEOPLE PRESENT #Absent
batch0 = ["Aaron","Weijie"] #Max 
batch1 = ["Jack", "Ivan"]
batch2 = ["Junyang", "Yicong", "Jowell", "Jonathan","Alvin"] 
batch3 = ["Bala", "Jinming", "Eugene", "Jian Yong"]
acf = ["Luke", "Ryan", "Stanley", "Yash"]
batch4 = ["Rayshawn",]
batch5 = ["Denver", "Praveen"]
stayout = ["Kaijie"]

#Off and Leave system #don't use if gone for whole mount. can just remove from array above
whoandwhenpresent = { #[name:day_of_mount_gone]
    'Aaron': [1,2],
}



workbook = Workbook()
sheet = workbook.active
sheet["A1"] = "TIME/NAME"
alphabets = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]

#initialise
noofdays = int(input("Number of days mounting: "))
if noofdays == 3:
    status = "weekend"
else:
    status = "weekday"

def assigning(row, duty):
    randomperson = random.randint(2, peoplepresent+1)
    if row == 2: #if first row just put only
        if sheet.cell(row = row, column = randomperson).value == None:
            sheet.cell(row = row, column = randomperson).value = duty
            sheet.cell(row = row + 1, column = randomperson).value = duty #second half of duty
        else:
            assigning(row, duty)
    else: #ensures they have rest before duty.
        if sheet.cell(row = row, column = randomperson).value == None and sheet.cell(row = row-1, column = randomperson).value == None:
            sheet.cell(row = row, column = randomperson).value = duty
            sheet.cell(row = row + 1, column = randomperson).value = duty #second half of duty
        else:
            assigning(row, duty)

def assigningpeak(row, duty): #4hourblock 1 peak 1 non-peak
    randomperson = random.randint(2, peoplepresent+1)
    if sheet.cell(row = row, column = randomperson).value == None and sheet.cell(row = row-1, column = randomperson).value == None:
        sheet.cell(row = row, column = randomperson).value = duty
    else:
        assigningpeak(row, duty)

def assigningafterpeak(counter,duty):
    if sheet.cell(row = i, column = counter).value == duty:
        sheet.cell(row = i+1, column = counter).value = duty
    else:
        counter += 1
        assigningafterpeak(counter,duty)

def countcellstoleft(row): #counter
    answer = 0
    for i in range(2, peoplepresent+2):
        if sheet.cell(row= row, column= i).value != None:
            answer += 1
    return answer

def hourscounter():
    for i in range(2, peoplepresent+2):
        counterhour = 0
        for row in range(2, totalrows):
            if sheet.cell(row=row, column = i).value != None:
                counterhour += 1
        sheet.cell(row=hoursrow, column = i).value = counterhour*2

def xinjiaolaojiaosystem():
    hoursranking = {}
    for i in range(2, peoplepresent+2):
        hours = sheet.cell(row=hoursrow, column = i).value
        hoursranking[i] = hours
    hoursranking = {k: v for k, v in sorted(hoursranking.items(), key=lambda item: item[1])}
    columnorder = list(hoursranking.keys()) 
    for i in range(len(team)-1):
        sheet.cell(row=1, column=columnorder[i]).value = team[i]
    return



#set up time
row = 2
timedefault = ["1100-1300", "1300-1500","1500-1700", "1700-1900","1900-2100", "2100-2300", "2300-0100","0100-0300", "0300-0500","0500-0700","0700-0900","0900-1100"]
times = noofdays * timedefault
for timeblock in times:
    sheet.cell(row = row, column = 1).value = timeblock
    row += 1
totalrows = row
hoursrow = totalrows #dunnid to add one more because final interation of timeblock already adds 1 more.
sheet.cell(row = hoursrow, column = 1).value = "TOTAL HOURS"

#set up humans, minimum 19
team = batch0 + batch1 + batch2 + batch3 + batch4 + batch5 + stayout + ["COUNTER"]
peoplepresent = len(team)-1
print("Number of people present is {} excluding 4 going to copper".format(str(peoplepresent))) 
column = 2
for name in team:
    sheet.cell(row = 1, column = column).value = name
    column += 1

def process_stayout_first():
    stayoutcolumn = len(team)
    for i in range(2, totalrows):
        if status == "weekday" and (sheet.cell(row=i, column=1).value in ["1700-1900","1900-2100","2100-2300","2300-0100","0100-0300","0300-0500","0500-0700"]):
            sheet.cell(row=i, column=stayoutcolumn).value = "STAYOUT"
        if status == "weekend":
            if i >= 5 and i <= totalrows-3:
                sheet.cell(row=i, column=stayoutcolumn).value = "STAYOUT"
process_stayout_first()

#def leaveandoffs(who,whichdays):
for name in team:
    if name in list(whoandwhenpresent.keys()):
        for daynumber in whoandwhenpresent[name]:
            print(name + " is on leave for day " + str(daynumber))
            if daynumber == 1:
                sheet.cell(row="person's row",column="person's column").value = "OFF/LL"

#colour coding
def colourthisrow(row,colour):
    for i in range(0,len(team)+1): #+1 cuz need account for time and counter column
        columncoordinate = alphabets[i]
        cellcoordinate = columncoordinate + str(row)
        cell = sheet[cellcoordinate]
        cell.fill = PatternFill(start_color=colour, end_color=colour, fill_type = "solid")


#initialise duties
non_peak = ["TG", "18", "XSVC", "XCBT", "CCTV", "CCTV2", "VACS"]
peak = ["TG", "18", "XSVC", "XCBT", "CCTV", "CCTV2", "VACS","TG2", "XSVC2", "CHKR"]
silent = [e for e in non_peak if e not in ('XSVC', 'XCBT')]

#assign dutytypes to hours
#nonpeak = 7, peak = 10, silent = 5
row = 2 #reset row again
print("planning....")

if status == "weekday":
    for i in range(2, totalrows):
        if i%2 == 0: #iterates across even rows only so that we assign duty every 4 hours
            if (sheet.cell(row= i, column = 1).value in ["1100-1300", "1300-1500","1500-1700","1700-1900","0900-1100"]): #if non_peak on normal hours
                for duty in non_peak:
                    assigning(i, duty)
                #if cell is empty (leave, off, MA etc) then put into random
            if (sheet.cell(row= i, column = 1).value in ["0700-0900"]):
                colourthisrow(i,"ff0000")
                for duty in peak:
                    assigningpeak(i,duty)
                counter = 1 #function below for adding non-peak for 0900-1100
                for duty in non_peak:
                    assigningafterpeak(counter,duty)
            if (sheet.cell(row= i, column = 1).value in ["1900-2100","2100-2300","2300-0100","0100-0300","0300-0500","0500-0700"]):
                colourthisrow(i,"808080")
                colourthisrow(i+1,"808080")
                for duty in silent:
                    assigning(i,duty)
        sheet.cell(row=i, column= peoplepresent+2).value = countcellstoleft(i)
elif status == "weekend":
    for i in range(2, totalrows):
        if i%2 == 0: #iterates across even rows only so that we assign duty every 4 hours
            if (sheet.cell(row= i, column = 1).value in ["1100-1300", "1300-1500","1500-1700","1700-1900","0900-1100"]): #if non_peak on normal hours
                if i<=4 or i>=35:
                    for duty in non_peak:
                        assigning(i, duty)
                    #if cell is empty (leave, off, MA etc) then put into random
                else:
                    colourthisrow(i,"808080")
                    colourthisrow(i+1,"808080")
                    for duty in silent:
                        assigning(i,duty)
            if (sheet.cell(row= i, column = 1).value in ["0700-0900"]):
                if i>=35:
                    colourthisrow(i,"ff0000")
                    for duty in peak:
                        assigningpeak(i,duty)
                    counter = 1 #function below for adding non-peak for 0900-1100
                    for duty in non_peak:
                        assigningafterpeak(counter,duty)
                else:
                    colourthisrow(i,"808080")
                    colourthisrow(i+1,"808080")
                    for duty in silent:
                        assigning(i,duty)
            if (sheet.cell(row= i, column = 1).value in ["1900-2100","2100-2300","2300-0100","0100-0300","0300-0500","0500-0700"]):
                colourthisrow(i,"808080")
                colourthisrow(i+1,"808080")
                for duty in silent:
                    assigning(i,duty)
        sheet.cell(row=i, column= peoplepresent+2).value = countcellstoleft(i)




hourscounter()
#xinjiaolaojiaosystem()
print("Done.")


workbook.save(filename=storagelocation)