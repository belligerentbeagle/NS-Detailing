import xlsxwriter
import os
import pandas as pd
from openpyxl import Workbook
import random

def assigning(row, duty):
    randomperson = random.randint(2, peoplepresent)
    if row == 2:
        if sheet.cell(row = row, column = randomperson).value == None:
            sheet.cell(row = row, column = randomperson).value = duty
        else:
            assigning(row, duty)
    else: #ensures they have rest before duty.
        if sheet.cell(row = row, column = randomperson).value == None and sheet.cell(row = row-1, column = randomperson).value == None:
            sheet.cell(row = row, column = randomperson).value = duty
        else:
            assigning(row, duty)

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "TIME/NAME"

#set up time
noofdays = int(input("Number of days mounting: "))
row = 2
timedefault = ["1100-1500", "1500-1900", "1900-2300", "2300-0300", "0300-0700", "0700-1100",]
times = noofdays * timedefault
for timeblock in times:
    sheet.cell(row = row, column = 1).value = timeblock
    row += 1
totalrows = row

#set up humans
batch1 = ["Jack", "Ivan"]
batch2 = ["Junyang", "Alvin", "yicong", "Jowell", "Jonathan"]
batch3 = ["Bala", "Jinming", "eugene", "Jian yong"]
acf = ["Luke", "Ryan", "Stanley", "yash"]
batch4 = ["Rayshawn", "Kaijie"]
batch5 = ["Denver", "Praveen"]
team = batch1 + batch2 + batch3 + batch4 + batch5 + ["COUNTER"]
peoplepresent = len(team)
column = 2
for name in team:
    sheet.cell(row = 1, column = column).value = name
    column += 1

#initialise duties
non_peak = ["TG", "18", "XSVC", "XCBT", "CCTV", "CCTV2", "VACS"]
peak = ["TG", "18", "XSVC", "XCBT", "CCTV", "CCTV2", "VACS","TG2", "XSVC2", "CHKR"]
silent = [e for e in non_peak if e not in ('XSVC', 'XCBT')]

#assign dutytypes to hours
#nonpeak = 7, peak = 10, silent = 5
row = 2 #reset row again
status = "weekday" #to be variable

for i in range(2, totalrows):
    if status == "weekday" and (sheet.cell(row= i, column = 1).value in ["1100-1500", "1500-1900", "0700-1100"]): #if non_peak on normal hours
        print("it is non-peak at " + sheet.cell(row= i, column = 1).value)
        for duty in non_peak:
            assigning(i, duty)
        #if cell is empty (leave, off, MA etc) then put into random
    else:
        print("it is not non-peak at " + sheet.cell(row= i, column = 1).value)







workbook.save(filename="/Users/weiyushit/OneDrive/Github stuff/NS Detailing/hello_world.xlsx")


'''
whereami = os.getcwd()
print(whereami)

#changes directory for my excel
os.chdir("/Users/weiyushit/OneDrive/Github stuff/NS Detailing")

# List all files and directories in current directory -> print(os.listdir('.'))

# Specify a writer
writer = pd.ExcelWriter('example.xlsx', engine='openpyxl')
file = 'example.xlsx'

# Load spreadsheet
xl = pd.ExcelFile(file)
'''
